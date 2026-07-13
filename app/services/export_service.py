import csv
import io
import re
import zipfile
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
import base64
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, Image


def clean_markdown(text: str) -> str:
    if not text:
        return ""
    # Remove bold/italic markup
    text = re.sub(r"\*\*([^*]+)\*\*", r"\1", text)
    text = re.sub(r"\*([^*]+)\*", r"\1", text)
    text = re.sub(r"__([^_]+)__", r"\1", text)
    text = re.sub(r"_([^_]+)_", r"\1", text)
    # Remove headers marker
    text = re.sub(r"#+\s+", "", text)
    return text


def generate_excel(
    title: str,
    headers: list[str],
    rows: list[dict],
    summary: str | None = None,
    kpis: list[dict] | None = None
) -> bytes:
    wb = Workbook()
    
    # Check if we should add a Summary tab
    if summary or kpis:
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_sum.views.sheetView[0].showGridLines = True
        
        # Style definition
        title_font = Font(name="Calibri", size=16, bold=True, color="1B365D")
        section_font = Font(name="Calibri", size=12, bold=True, color="1B365D")
        label_font = Font(name="Calibri", size=11, bold=True)
        value_font = Font(name="Calibri", size=11)
        
        # Title
        ws_sum.cell(row=1, column=1, value=title).font = title_font
        ws_sum.row_dimensions[1].height = 30
        
        row_idx = 3
        if kpis:
            ws_sum.cell(row=row_idx, column=1, value="Key Performance Indicators").font = section_font
            ws_sum.row_dimensions[row_idx].height = 20
            row_idx += 1
            
            # Header for KPIs
            ws_sum.cell(row=row_idx, column=1, value="Metric").font = label_font
            ws_sum.cell(row=row_idx, column=2, value="Value").font = label_font
            ws_sum.cell(row=row_idx, column=3, value="Status").font = label_font
            row_idx += 1
            
            for kpi in kpis:
                ws_sum.cell(row=row_idx, column=1, value=kpi.get("label", "")).font = value_font
                ws_sum.cell(row=row_idx, column=2, value=kpi.get("value", "")).font = value_font
                ws_sum.cell(row=row_idx, column=3, value=kpi.get("change", "")).font = value_font
                row_idx += 1
                
            row_idx += 1  # Spacer
            
        if summary:
            ws_sum.cell(row=row_idx, column=1, value="AI Insights & Summary").font = section_font
            ws_sum.row_dimensions[row_idx].height = 20
            row_idx += 1
            
            cleaned_text = clean_markdown(summary)
            # Write lines
            for line in cleaned_text.split("\n"):
                if line.strip():
                    ws_sum.cell(row=row_idx, column=1, value=line.strip()).font = value_font
                    row_idx += 1
                    
        # Auto-adjust Summary columns
        for col in ws_sum.columns:
            max_len = 0
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col[0].column)
            ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 15)
            
        ws = wb.create_sheet(title="Data")
    else:
        ws = wb.active
        ws.title = "Report"

    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True

    # Style definitions
    title_font = Font(name="Calibri", size=16, bold=True, color="1B365D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")

    data_font = Font(name="Calibri", size=11)
    zebra_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")

    thin_side = Side(border_style="thin", color="E0E0E0")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Title
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40

    # Empty row
    ws.append([])
    ws.row_dimensions[2].height = 15

    # Headers
    ws.append(headers)
    header_row_idx = 3
    ws.row_dimensions[header_row_idx].height = 28

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = cell_border

    # Data Rows
    for r_idx, row in enumerate(rows, start=4):
        row_values = [row.get(h, "") for h in headers]
        ws.append(row_values)
        ws.row_dimensions[r_idx].height = 20

        is_even = (r_idx % 2 == 0)
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = cell_border
            if is_even:
                cell.fill = zebra_fill

            val = cell.value
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(val, float):
                    cell.number_format = '0.00'
                else:
                    cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[2:]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_pdf(
    title: str,
    headers: list[str],
    rows: list[dict],
    summary: str | None = None,
    chart_image: str | None = None,
    kpis: list[dict] | None = None
) -> bytes:
    # Use landscape letter if there are more than 6 columns, else portrait
    page_size = landscape(letter) if len(headers) > 6 else letter

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1B365D'),
        spaceAfter=15
    )

    header_style = ParagraphStyle(
        'TableHeader',
        fontSize=9,
        leading=11,
        fontName='Helvetica-Bold',
        textColor=colors.white
    )

    body_style = ParagraphStyle(
        'TableBody',
        fontSize=8,
        leading=10,
        fontName='Helvetica',
        textColor=colors.HexColor('#333333')
    )

    elements = []
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 10))

    # Add KPIs if provided
    if kpis:
        kpi_data = []
        kpi_row = []
        for kpi in kpis:
            label = kpi.get("label", "")
            val = kpi.get("value", "")
            change = kpi.get("change", "")
            cell_content = f"<b>{val}</b><br/><font color='#555555' size='8'>{label}</font>"
            kpi_style = ParagraphStyle(
                'KPIStyle',
                fontSize=12,
                leading=14,
                fontName='Helvetica',
                alignment=1
            )
            kpi_row.append(Paragraph(cell_content, kpi_style))
        kpi_data.append(kpi_row)
        
        kpi_col_width = doc.width / len(kpis)
        kpi_table = Table(kpi_data, colWidths=[kpi_col_width] * len(kpis))
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0F4FA')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#D0DCEF')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 15))

    # Add Chart Image if provided
    if chart_image:
        try:
            base64_str = chart_image
            if "," in base64_str:
                base64_str = base64_str.split(",", 1)[1]
            img_data = base64.b64decode(base64_str)
            img_buf = io.BytesIO(img_data)
            
            chart_w = doc.width * 0.95
            chart_h = chart_w * 0.5
            
            img_flowable = Image(img_buf, width=chart_w, height=chart_h)
            elements.append(img_flowable)
            elements.append(Spacer(1, 15))
        except Exception as e:
            warning_style = ParagraphStyle(
                'WarningStyle',
                fontSize=9,
                textColor=colors.HexColor('#CC0000')
            )
            elements.append(Paragraph(f"Could not render chart: {str(e)}", warning_style))
            elements.append(Spacer(1, 10))

    # Add Summary if provided
    if summary:
        summary_title_style = ParagraphStyle(
            'SummaryTitle',
            parent=styles['Heading2'],
            fontSize=12,
            leading=15,
            textColor=colors.HexColor('#1B365D'),
            spaceAfter=6
        )
        summary_body_style = ParagraphStyle(
            'SummaryBody',
            fontSize=9,
            leading=13,
            fontName='Helvetica',
            textColor=colors.HexColor('#333333'),
            spaceAfter=15
        )
        elements.append(Paragraph("AI Insights & Executive Summary", summary_title_style))
        
        cleaned_text = clean_markdown(summary)
        for paragraph_text in cleaned_text.split("\n\n"):
            if paragraph_text.strip():
                clean_p = paragraph_text.replace("\n", " ").strip()
                elements.append(Paragraph(clean_p, summary_body_style))

    # Add Data Table if provided and contains rows
    if rows and headers:
        # Separate table onto a new page if we already added a summary/chart
        if summary or chart_image or kpis:
            elements.append(PageBreak())

        # Define right-aligned numeric body style
        right_body_style = ParagraphStyle(
            'TableBodyRight',
            parent=body_style,
            alignment=2  # Right-aligned
        )

        table_data = []
        header_row = [Paragraph(h, header_style) for h in headers]
        table_data.append(header_row)

        for row in rows:
            row_cells = []
            for h in headers:
                val = row.get(h, "")
                is_num = isinstance(val, (int, float, Decimal))
                if val is None:
                    val_str = ""
                elif isinstance(val, float):
                    val_str = f"{val:.2f}"
                elif isinstance(val, int):
                    val_str = f"{val:,}"
                elif isinstance(val, Decimal):
                    val_str = f"{float(val):.2f}"
                else:
                    val_str = str(val)
                
                style = right_body_style if is_num else body_style
                row_cells.append(Paragraph(val_str, style))
            table_data.append(row_cells)

        doc_width = doc.width
        
        # Calculate dynamic column widths based on content lengths
        col_widths = []
        for h in headers:
            max_len = len(h)
            for row in rows:
                val = row.get(h, "")
                if val is None:
                    val_str = ""
                elif isinstance(val, float):
                    val_str = f"{val:.2f}"
                elif isinstance(val, int):
                    val_str = f"{val:,}"
                elif isinstance(val, Decimal):
                    val_str = f"{float(val):.2f}"
                else:
                    val_str = str(val)
                max_len = max(max_len, len(val_str))
            
            # Convert character count to point width, minimum 50, maximum 250
            approx_width = min(max(max_len * 6.5, 50.0), 250.0)
            col_widths.append(approx_width)
            
        # Scale proportionally to fit page width
        total_width = sum(col_widths)
        scale_factor = doc_width / total_width
        col_widths = [w * scale_factor for w in col_widths]

        t = Table(table_data, colWidths=col_widths, repeatRows=1)

        t_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B365D')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ])

        for i in range(1, len(rows) + 1):
            if i % 2 == 0:
                t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F7F9FC'))

        t.setStyle(t_style)
        elements.append(t)

    doc.build(elements)
    return buffer.getvalue()


def generate_bulk_excel(reports: list[dict]) -> bytes:
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    for idx, rep in enumerate(reports):
        title = rep.get("title", f"Report {idx+1}")
        headers = rep.get("headers", [])
        rows = rep.get("rows", [])

        sheet_title = re.sub(r"[\\*?:/\[\]]", "", title)[:30]
        if not sheet_title.strip():
            sheet_title = f"Report {idx+1}"

        base_title = sheet_title
        counter = 1
        while sheet_title in wb.sheetnames:
            suffix = f" {counter}"
            sheet_title = f"{base_title[:30-len(suffix)]}{suffix}"
            counter += 1

        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True

        title_font = Font(name="Calibri", size=16, bold=True, color="1B365D")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")

        data_font = Font(name="Calibri", size=11)
        zebra_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")

        thin_side = Side(border_style="thin", color="E0E0E0")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        ws.append([])
        ws.row_dimensions[2].height = 15

        ws.append(headers)
        header_row_idx = 3
        ws.row_dimensions[header_row_idx].height = 28

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = cell_border

        for r_idx, row in enumerate(rows, start=4):
            row_values = [row.get(h, "") for h in headers]
            ws.append(row_values)
            ws.row_dimensions[r_idx].height = 20

            is_even = (r_idx % 2 == 0)
            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = data_font
                cell.border = cell_border
                if is_even:
                    cell.fill = zebra_fill

                val = cell.value
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(val, float):
                        cell.number_format = '0.00'
                    else:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col[2:]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_bulk_pdf(reports: list[dict]) -> bytes:
    has_large_report = any(len(rep.get("headers", [])) > 6 for rep in reports)
    page_size = landscape(letter) if has_large_report else letter

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1B365D'),
        spaceAfter=15
    )

    header_style = ParagraphStyle(
        'TableHeader',
        fontSize=9,
        leading=11,
        fontName='Helvetica-Bold',
        textColor=colors.white
    )

    body_style = ParagraphStyle(
        'TableBody',
        fontSize=8,
        leading=10,
        fontName='Helvetica',
        textColor=colors.HexColor('#333333')
    )

    elements = []

    for idx, rep in enumerate(reports):
        title = rep.get("title", f"Report {idx+1}")
        headers = rep.get("headers", [])
        rows = rep.get("rows", [])

        if idx > 0:
            elements.append(PageBreak())

        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 10))

        table_data = []
        header_row = [Paragraph(h, header_style) for h in headers]
        table_data.append(header_row)

        for row in rows:
            row_cells = []
            for h in headers:
                val = row.get(h, "")
                if val is None:
                    val_str = ""
                elif isinstance(val, float):
                    val_str = f"{val:.2f}"
                else:
                    val_str = str(val)
                row_cells.append(Paragraph(val_str, body_style))
            table_data.append(row_cells)

        doc_width = doc.width
        col_width = doc_width / max(len(headers), 1)
        col_widths = [col_width] * len(headers)

        t = Table(table_data, colWidths=col_widths, repeatRows=1)

        t_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B365D')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
        ])

        for i in range(1, len(rows) + 1):
            if i % 2 == 0:
                t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F7F9FC'))

        t.setStyle(t_style)
        elements.append(t)

    doc.build(elements)
    return buffer.getvalue()


def generate_bulk_zip(reports: list[dict], format_type: str) -> bytes:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for idx, rep in enumerate(reports):
            title = rep.get("title", f"Report_{idx+1}")
            headers = rep.get("headers", [])
            rows = rep.get("rows", [])

            safe_title = re.sub(r"[^\w\-_]", "_", title)

            if format_type == "csv":
                csv_buffer = io.StringIO()
                writer = csv.writer(csv_buffer)
                writer.writerow(headers)
                for row in rows:
                    writer.writerow([row.get(h, "") for h in headers])
                zip_file.writestr(f"{safe_title}.csv", csv_buffer.getvalue())
            elif format_type == "excel":
                excel_bytes = generate_excel(title, headers, rows)
                zip_file.writestr(f"{safe_title}.xlsx", excel_bytes)
            elif format_type == "pdf":
                pdf_bytes = generate_pdf(title, headers, rows)
                zip_file.writestr(f"{safe_title}.pdf", pdf_bytes)

    return zip_buffer.getvalue()
